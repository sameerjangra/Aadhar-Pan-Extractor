import pandas as pd
import os

from typing import List, Dict, Any

def export_to_excel(data: List[Dict[str, Any]], output_path: str = "extraction_results.xlsx"):
    """
    Save the extracted list of dictionaries to an Excel file with formatting.
    """
    try:
        # Create a DataFrame directly from the list of dicts
        df = pd.DataFrame(data)
        
        # Ensure the directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # Use ExcelWriter with valid engine
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            
            # Find "Photo Path" column index (1-based for openpyxl)
            photo_col_idx = None
            if "Photo Path" in df.columns:
                photo_col_idx = df.columns.get_loc("Photo Path") + 1

            # Iterate through columns to adjust width and enable wrap_text
            for idx, col in enumerate(df.columns):
                # Determine max length of data in column (header vs content)
                # We set a max width to prevent extremely wide columns
                max_len =  max((
                    df[col].astype(str).map(len).max(),
                    len(str(col))
                )) + 2 # Adding a little padding
                
                # Cap the width at 50 to ensure readability, longer text will wrap
                final_width = min(max_len, 50)
                
                # Column letter (A=1, B=2...)
                col_letter = chr(65 + idx) if idx < 26 else 'A' + chr(65 + (idx - 26)) # Simple logic for A-Z
                # Actually openpyxl has get_column_letter utils but let's just use 1-based index access or direct iteration
                
                cell = worksheet.cell(row=1, column=idx+1)
                column_letter = cell.column_letter
                
                worksheet.column_dimensions[column_letter].width = final_width
                
                # Apply word wrap to all cells in this column
                for row_idx in range(len(df) + 1): # +1 for header
                     cell = worksheet.cell(row=row_idx+1, column=idx+1)
                     cell.alignment = cell.alignment.copy(wrap_text=True)

            # Insert Images if column exists
            if photo_col_idx:
                from openpyxl.drawing.image import Image as OpenpyxlImage
                
                # Increase column width for Photo
                cell = worksheet.cell(row=1, column=photo_col_idx)
                worksheet.column_dimensions[cell.column_letter].width = 15
                
                for i, row in df.iterrows():
                    photo_path = row.get("Photo Path")
                    if photo_path and isinstance(photo_path, str) and os.path.exists(photo_path):
                        try:
                            img = OpenpyxlImage(photo_path)
                            # Resize to fit cell roughly, maintaining 35:45 aspect ratio (7:9)
                            # Let's use Height ~ 90 pixels (approx 24mm on screen? Excel units are weird).
                            # 35mm : 45mm -> 0.777 ratio.
                            img.height = 90
                            img.width = 70
                            
                            # Excel row is i+2 (1 for header, 1 for 0-index)
                            excel_row = i + 2
                            
                            # Set row height to accommodate image (Height in points. 1 px ~ 0.75 points?)
                            # 90 px is roughly 68 points.
                            worksheet.row_dimensions[excel_row].height = 75
                            
                            # Add image to cell
                            worksheet.add_image(img, cell.column_letter + str(excel_row))
                            
                            # Optional: Clear the text path so it looks clean? 
                            # worksheet.cell(row=excel_row, column=photo_col_idx).value = "" 
                        except Exception as e:
                            print(f"Error embedding image {photo_path}: {e}")

        print(f"Data successfully exported to {output_path}")
        return output_path
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        raise e
 